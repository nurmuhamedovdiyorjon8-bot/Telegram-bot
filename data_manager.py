import os
import json
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ===================== FAYL YO'LLARI =====================
DATA_DIR = "data"
ORDERS_FILE = os.path.join(DATA_DIR, "orders.xlsx")
CUSTOMERS_FILE = os.path.join(DATA_DIR, "customers.xlsx")
PRODUCTS_FILE = os.path.join(DATA_DIR, "products.xlsx")
SETTINGS_FILE = os.path.join(DATA_DIR, "settings.json")

os.makedirs(DATA_DIR, exist_ok=True)

# ===================== SOZLAMALAR =====================
def get_settings():
    if not os.path.exists(SETTINGS_FILE):
        default = {
            "location": "Samarqand, O'zbekiston",
            "work_start": "08:00",
            "work_end": "17:00",
            "phone": "+998 77 285 01 10",
            "min_order": 45000,
            "free_delivery_from": 100000,
            "welcome_text": "💧 Salom! Markiz Premium botiga xush kelibsiz!"
        }
        save_settings(default)
        return default
    with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def save_settings(settings):
    with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
        json.dump(settings, f, ensure_ascii=False, indent=2)

# ===================== MAHSULOTLAR =====================
def _init_products():
    wb = Workbook()
    ws = wb.active
    ws.title = "Mahsulotlar"
    headers = ["ID", "Nomi", "Narxi", "Rasm", "Kategoriya", "Mavjud"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0EA5E9")
    default_products = [
        [1, "Suv 19 L", 15000, "💧", "Suv", True],
        [2, "Suv 10 L", 8500, "💧", "Suv", True],
        [3, "Suv 5 L", 5500, "💧", "Suv", True],
        [4, "Pompa", 65000, "⚙️", "Aksessuar", True],
        [5, "19 L Dastacha", 55000, "🔧", "Aksessuar", True],
    ]
    for row in default_products:
        ws.append(row)
    wb.save(PRODUCTS_FILE)

def get_products():
    if not os.path.exists(PRODUCTS_FILE):
        _init_products()
    wb = load_workbook(PRODUCTS_FILE)
    ws = wb.active
    products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        products.append({
            "id": row[0],
            "name": row[1] or "",
            "price": float(row[2]) if row[2] else 0,
            "image": row[3] or "💧",
            "category": row[4] or "Boshqa",
            "available": row[5] if row[5] is not None else True
        })
    return [p for p in products if p.get("available")]

def update_product_price(product_id, new_price):
    if not os.path.exists(PRODUCTS_FILE):
        _init_products()
    wb = load_workbook(PRODUCTS_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == product_id:
            row[2].value = new_price
            break
    wb.save(PRODUCTS_FILE)

# ===================== MIJOZLAR =====================
def _ensure_customers_file():
    if not os.path.exists(CUSTOMERS_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Mijozlar"
        headers = ["Telegram ID", "Ism", "Telefon", "Buyurtmalar soni", "Jami summa", "Ro'yxatdan o'tgan", "So'nggi buyurtma"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0EA5E9")
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 16
        wb.save(CUSTOMERS_FILE)

def get_customers():
    _ensure_customers_file()
    wb = load_workbook(CUSTOMERS_FILE)
    ws = wb.active
    customers = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        customers.append({
            "telegram_id": row[0],
            "name": row[1] or "",
            "phone": row[2] or "",
            "orders_count": row[3] or 0,
            "total_spent": row[4] or 0,
            "registered": row[5] or "",
            "last_order": row[6] or "",
        })
    return customers

def get_or_create_customer(telegram_id, name, phone=""):
    _ensure_customers_file()
    customers = get_customers()
    for c in customers:
        if c['telegram_id'] == telegram_id:
            return c
    # Yangi mijoz qo'shish
    wb = load_workbook(CUSTOMERS_FILE)
    ws = wb.active
    ws.append([telegram_id, name, phone, 0, 0, datetime.now().strftime("%Y-%m-%d %H:%M"), ""])
    wb.save(CUSTOMERS_FILE)
    return {"telegram_id": telegram_id, "name": name, "phone": phone, "orders_count": 0, "total_spent": 0}

def update_customer_phone(telegram_id, phone):
    _ensure_customers_file()
    wb = load_workbook(CUSTOMERS_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == telegram_id:
            row[2].value = phone
            break
    wb.save(CUSTOMERS_FILE)

def increment_customer_orders(telegram_id, order_total):
    _ensure_customers_file()
    wb = load_workbook(CUSTOMERS_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == telegram_id:
            row[3].value = (row[3].value or 0) + 1
            row[4].value = (row[4].value or 0) + order_total
            row[6].value = datetime.now().strftime("%Y-%m-%d %H:%M")
            break
    wb.save(CUSTOMERS_FILE)

# ===================== BUYURTMALAR =====================
def _ensure_orders_file():
    if not os.path.exists(ORDERS_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Buyurtmalar"
        headers = ["ID", "Mijoz ID", "Mijoz ismi", "Telefon", "Manzil", "Qavat", "To'lov", "Mahsulotlar", "Jami summa", "Holat", "Sana"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0EA5E9")
        for col_letter, width in [('C', 20), ('D', 16), ('E', 30), ('H', 50)]:
            ws.column_dimensions[col_letter].width = width
        wb.save(ORDERS_FILE)

def _get_next_order_id():
    _ensure_orders_file()
    wb = load_workbook(ORDERS_FILE)
    ws = wb.active
    max_id = 1000
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and isinstance(row[0], int):
            max_id = max(max_id, row[0])
    return max_id + 1

def save_order(customer_id, customer_name, items, address, floor, payment, phone):
    _ensure_orders_file()
    order_id = _get_next_order_id()
    total = sum(i['price'] * i['qty'] for i in items)
    items_text = "; ".join([f"{i['name']} x{i['qty']} ({i['price']*i['qty']:,} so'm)" for i in items])
    date_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    wb = load_workbook(ORDERS_FILE)
    ws = wb.active
    ws.append([order_id, customer_id, customer_name, phone, address, floor, payment, items_text, total, "Yangi", date_str])
    wb.save(ORDERS_FILE)

    # Mijoz statistikasini yangilash
    increment_customer_orders(customer_id, total)

    return {
        "id": order_id,
        "customer_id": customer_id,
        "customer_name": customer_name,
        "phone": phone,
        "address": address,
        "floor": floor,
        "payment": payment,
        "items": items,
        "total": total,
        "status": "Yangi",
        "date": date_str
    }

def get_orders():
    _ensure_orders_file()
    wb = load_workbook(ORDERS_FILE)
    ws = wb.active
    orders = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        # items ni parse qilish (oddiy text formatda)
        items_raw = row[7] or ""
        items = []
        for part in items_raw.split(";"):
            part = part.strip()
            if part:
                items.append({"name": part, "qty": 1, "price": 0, "id": 0})
        orders.append({
            "id": row[0],
            "customer_id": row[1],
            "customer_name": row[2] or "",
            "phone": row[3] or "",
            "address": row[4] or "",
            "floor": row[5] or "",
            "payment": row[6] or "",
            "items": items,
            "total": row[8] or 0,
            "status": row[9] or "Yangi",
            "date": str(row[10]) if row[10] else "",
        })
    return orders

def get_customer_orders(customer_id):
    all_orders = get_orders()
    return [o for o in all_orders if o['customer_id'] == customer_id]

def update_order_status(order_id, new_status):
    _ensure_orders_file()
    wb = load_workbook(ORDERS_FILE)
    ws = wb.active
    found_order = None
    for row in ws.iter_rows(min_row=2):
        if row[0].value == order_id:
            row[9].value = new_status
            found_order = {
                "id": order_id,
                "customer_id": row[1].value,
                "status": new_status
            }
            break
    if found_order:
        wb.save(ORDERS_FILE)
    return found_order
